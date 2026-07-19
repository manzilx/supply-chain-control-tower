"""Universal ingestion engine — drop any Excel workbook or CSV, get data in.

The pipeline:

  1. **Parse** — every sheet (xlsx) or the single table (csv) into header+rows.
  2. **Classify** — what entity does each sheet hold? Sheet-name hints first,
     then header-signature scoring (how many canonical fields match).
  3. **Map columns** — normalise headers, exact synonym lookup, then fuzzy
     (difflib). Anything still unmapped is optionally sent to Grok in ONE
     batched call; deterministic behaviour is unchanged when no key is set.
  4. **Validate + coerce** — per-entity row validation with row-numbered
     errors (same spirit as the BOM CSV uploader, generalised).
  5. **Stage** — preview returns everything (mapping, errors, samples) plus a
     staging_id; nothing touches the stores yet.
  6. **Commit** — replays the staged rows into planning / vendor_store with
     audit events, then busts the analytics caches.

Supported entities: projects, BOM lines, suppliers.
"""

from __future__ import annotations

import csv
import difflib
import io
import re
import time
import uuid
from datetime import date, datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from .schemas import (
    BOMItem,
    IngestCommitReply,
    IngestPreviewReply,
    IngestSheetPreview,
    Milestone,
    Project,
    SupplierRecord,
    User,
)

# --- Canonical fields + synonyms ---------------------------------------------

# entity -> canonical field -> synonyms (normalised: lowercase, alnum+space)
SYNONYMS: Dict[str, Dict[str, List[str]]] = {
    "bom": {
        "code": ["code", "part no", "part number", "item code", "material", "material code", "sku", "part"],
        "description": ["description", "desc", "item", "item description", "material description", "name"],
        "quantity": ["quantity", "qty", "amount", "count"],
        "uom": ["uom", "unit", "units", "unit of measure"],
        "unit_cost_usd": ["unit cost usd", "unit cost", "unit price", "price", "cost", "rate", "usd"],
        "supplier_name": ["supplier name", "supplier", "vendor", "preferred supplier", "manufacturer"],
        "spec_doc_id": ["spec doc id", "spec", "specification", "spec ref"],
        "drawing_id": ["drawing id", "drawing", "dwg", "drawing ref"],
        "long_lead_days": ["long lead days", "lead time", "lead time days", "lead", "lead days"],
        "planned_need_date": ["planned need date", "need by", "need date", "required date", "ros", "required on site", "delivery date"],
        "milestone_code": ["milestone code", "milestone", "ms"],
        "project_id": ["project id", "project"],
        "status": ["status", "state"],
        "parent_item_id": ["parent item id", "parent"],
        "bom_item_id": ["bom item id", "item id", "line id", "line no"],
    },
    "project": {
        "project_id": ["project id", "id", "project code", "code"],
        "name": ["name", "project name", "title", "project"],
        "client": ["client", "customer", "owner", "employer"],
        "site": ["site", "location", "place"],
        "sector": ["sector", "industry", "type", "segment"],
        "start_date": ["start date", "start", "commencement"],
    },
    "supplier": {
        "name": ["name", "supplier", "vendor", "company", "supplier name", "vendor name"],
        "category": ["category", "commodity", "material category", "scope"],
        "country": ["country", "origin", "location"],
        "lead_time_days": ["lead time days", "lead time", "lead"],
        "on_time_delivery_pct": ["on time delivery pct", "otd", "otd pct", "on time", "on time pct", "otd %", "on time %"],
        "quality_ppm": ["quality ppm", "ppm", "defects ppm", "defect ppm"],
        "annual_spend_usd": ["annual spend usd", "annual spend", "spend", "spend usd"],
        "approved_alternatives": ["approved alternatives", "alternates", "alternatives", "alt count"],
        "risk_flags": ["risk flags", "flags", "risks"],
    },
}

REQUIRED: Dict[str, List[str]] = {
    "bom": ["code", "description", "quantity"],
    "project": ["project_id", "name"],
    "supplier": ["name", "category", "country"],
}

SHEET_NAME_HINTS = [
    ("project", re.compile(r"project", re.I)),
    ("supplier", re.compile(r"supplier|vendor", re.I)),
    ("bom", re.compile(r"bom|material|item|line", re.I)),
]

_VALID_BOM_STATUSES = {"spec_missing", "planned", "requisitioned", "ordered", "delivered"}


def _norm(header: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(header).strip().lower()).strip()


# --- Parsing -------------------------------------------------------------------


def _parse_csv(data: bytes) -> List[Tuple[str, List[str], List[List[Any]]]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [r for r in reader if any(str(c).strip() for c in r)]
    if not rows:
        return []
    return [("csv", [str(h).strip() for h in rows[0]], rows[1:])]


def _parse_xlsx(data: bytes) -> List[Tuple[str, List[str], List[List[Any]]]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheets: List[Tuple[str, List[str], List[List[Any]]]] = []
    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        headers: List[str] = []
        body: List[List[Any]] = []
        for raw in rows_iter:
            if not headers:
                cells = [str(c).strip() if c is not None else "" for c in raw]
                if sum(1 for c in cells if c) >= 2:  # first row with ≥2 filled cells = header
                    headers = cells
                continue
            if raw is None or all(c is None or str(c).strip() == "" for c in raw):
                continue
            body.append(list(raw))
        if headers and body:
            sheets.append((ws.title, headers, body))
    wb.close()
    return sheets


# --- Classification + mapping ----------------------------------------------


def _classify(sheet_name: str, headers: List[str]) -> Optional[str]:
    for entity, pattern in SHEET_NAME_HINTS:
        if pattern.search(sheet_name):
            return entity
    # Header-signature scoring: entity with most canonical matches wins.
    normed = [_norm(h) for h in headers if h]
    best, best_score = None, 0
    for entity, fields in SYNONYMS.items():
        score = 0
        for canon, syns in fields.items():
            if any(h in syns for h in normed):
                score += 1
        if score > best_score:
            best, best_score = entity, score
    return best if best_score >= 2 else None


def _map_headers(entity: str, headers: List[str]) -> Tuple[Dict[str, int], List[str]]:
    """Map canonical field -> column index. Returns (mapping, unmapped_headers)."""
    fields = SYNONYMS[entity]
    mapping: Dict[str, int] = {}
    used: set[int] = set()
    normed = [(_norm(h), i) for i, h in enumerate(headers)]

    # Pass 1: exact synonym match
    for canon, syns in fields.items():
        for h, i in normed:
            if i in used or not h:
                continue
            if h in syns:
                mapping[canon] = i
                used.add(i)
                break
    # Pass 2: fuzzy match on the remainder
    for canon, syns in fields.items():
        if canon in mapping:
            continue
        for h, i in normed:
            if i in used or not h:
                continue
            if difflib.get_close_matches(h, syns, n=1, cutoff=0.86):
                mapping[canon] = i
                used.add(i)
                break

    unmapped = [headers[i] for _, i in normed if i not in used and headers[i]]
    return mapping, unmapped


def _llm_assist_mapping(entity: str, unmapped: List[str], mapping: Dict[str, int], headers: List[str]) -> Dict[str, int]:
    """One batched Grok call to map leftover columns. No-op without a key."""
    from .llm import grok_json, is_enabled
    if not unmapped or not is_enabled():
        return mapping
    missing_fields = [f for f in SYNONYMS[entity] if f not in mapping]
    if not missing_fields:
        return mapping
    result = grok_json(
        "You map spreadsheet column headers to canonical schema fields for a "
        "procurement system. Reply with a JSON object whose keys are the given "
        "headers and values are one of the canonical field names or null.",
        f"Canonical fields: {missing_fields}\nHeaders to map: {unmapped}",
        max_tokens=300,
        timeout=12,
    )
    if not isinstance(result, dict):
        return mapping
    for header, canon in result.items():
        if canon in missing_fields and header in headers and canon not in mapping:
            mapping[canon] = headers.index(header)
    return mapping


# --- Row coercion -------------------------------------------------------------


def _cell(row: List[Any], mapping: Dict[str, int], field: str) -> Optional[str]:
    idx = mapping.get(field)
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _to_float(s: Optional[str]) -> Optional[float]:
    if s is None:
        return None
    cleaned = re.sub(r"[,$\s]", "", s)
    try:
        return float(cleaned)
    except ValueError:
        return None


def _to_int(s: Optional[str]) -> Optional[int]:
    f = _to_float(s)
    return int(f) if f is not None else None


def _to_date(s: Optional[str]) -> Optional[date]:
    if s is None:
        return None
    s = s.split(" ")[0].split("T")[0]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _validate_rows(entity: str, mapping: Dict[str, int], rows: List[List[Any]]) -> Tuple[List[dict], List[str]]:
    """Coerce rows into normalised dicts; collect row-numbered errors."""
    out: List[dict] = []
    errors: List[str] = []
    for idx, raw in enumerate(rows, start=2):  # row 1 = header
        if entity == "bom":
            code = _cell(raw, mapping, "code")
            desc = _cell(raw, mapping, "description")
            qty = _to_float(_cell(raw, mapping, "quantity"))
            if not code or not desc or qty is None:
                errors.append(f"row {idx}: missing code/description/quantity")
                continue
            if qty <= 0:
                errors.append(f"row {idx}: quantity must be positive")
                continue
            status = (_cell(raw, mapping, "status") or "").lower()
            spec = _cell(raw, mapping, "spec_doc_id")
            if status and status not in _VALID_BOM_STATUSES:
                errors.append(f"row {idx}: unknown status '{status}' — using default")
                status = ""
            out.append({
                "code": code, "description": desc, "quantity": qty,
                "uom": _cell(raw, mapping, "uom") or "EA",
                "unit_cost_usd": _to_float(_cell(raw, mapping, "unit_cost_usd")),
                "supplier_name": _cell(raw, mapping, "supplier_name"),
                "spec_doc_id": spec,
                "drawing_id": _cell(raw, mapping, "drawing_id"),
                "long_lead_days": _to_int(_cell(raw, mapping, "long_lead_days")),
                "planned_need_date": _to_date(_cell(raw, mapping, "planned_need_date")),
                "milestone_code": _cell(raw, mapping, "milestone_code"),
                "project_id": _cell(raw, mapping, "project_id"),
                "bom_item_id": _cell(raw, mapping, "bom_item_id"),
                "parent_item_id": _cell(raw, mapping, "parent_item_id"),
                "status": status or ("spec_missing" if not spec else "planned"),
            })
        elif entity == "project":
            pid = _cell(raw, mapping, "project_id")
            name = _cell(raw, mapping, "name")
            if not pid or not name:
                errors.append(f"row {idx}: missing project_id/name")
                continue
            out.append({
                "project_id": pid.upper().replace(" ", "-"),
                "name": name,
                "client": _cell(raw, mapping, "client") or "—",
                "site": _cell(raw, mapping, "site") or "—",
                "sector": _cell(raw, mapping, "sector") or "Industrial EPC",
                "start_date": _to_date(_cell(raw, mapping, "start_date")) or date.today(),
            })
        elif entity == "supplier":
            name = _cell(raw, mapping, "name")
            cat = _cell(raw, mapping, "category")
            country = _cell(raw, mapping, "country")
            if not name or not cat or not country:
                errors.append(f"row {idx}: missing name/category/country")
                continue
            flags_raw = _cell(raw, mapping, "risk_flags") or ""
            out.append({
                "name": name, "category": cat, "country": country,
                "lead_time_days": _to_int(_cell(raw, mapping, "lead_time_days")) or 60,
                "on_time_delivery_pct": _to_float(_cell(raw, mapping, "on_time_delivery_pct")) or 90.0,
                "quality_ppm": _to_int(_cell(raw, mapping, "quality_ppm")) or 500,
                "annual_spend_usd": _to_float(_cell(raw, mapping, "annual_spend_usd")) or 0.0,
                "approved_alternatives": _to_int(_cell(raw, mapping, "approved_alternatives")) or 0,
                "risk_flags": [f.strip() for f in re.split(r"[;,]", flags_raw) if f.strip()],
            })
    return out, errors


# --- Staging -------------------------------------------------------------------

_STAGING: Dict[str, dict] = {}
_STAGING_TTL_S = 1800  # 30 min


def _gc_staging() -> None:
    now = time.monotonic()
    for sid in [s for s, v in _STAGING.items() if now - v["at"] > _STAGING_TTL_S]:
        del _STAGING[sid]


# --- Public API -----------------------------------------------------------------


def preview(filename: str, data: bytes, user: User) -> IngestPreviewReply:
    _gc_staging()
    lower = filename.lower()
    if lower.endswith((".xlsx", ".xlsm")):
        parsed = _parse_xlsx(data)
    elif lower.endswith((".csv", ".txt")):
        parsed = _parse_csv(data)
    else:
        raise ValueError("Unsupported file type — upload .xlsx or .csv")

    sheets: List[IngestSheetPreview] = []
    staged_sheets: List[dict] = []
    for sheet_name, headers, rows in parsed:
        entity = _classify(sheet_name, headers)
        if entity is None:
            sheets.append(IngestSheetPreview(
                sheet=sheet_name, entity=None, rows_total=len(rows), rows_valid=0,
                mapped={}, unmapped=[h for h in headers if h],
                errors=["Could not classify this sheet — name it 'Projects', 'BOM' or 'Suppliers', or fix the headers."],
                sample=[],
            ))
            continue
        mapping, unmapped = _map_headers(entity, headers)
        mapping = _llm_assist_mapping(entity, unmapped, mapping, headers)
        unmapped = [h for i, h in enumerate(headers) if h and i not in mapping.values()]
        missing_required = [f for f in REQUIRED[entity] if f not in mapping]
        valid_rows: List[dict] = []
        errors: List[str] = []
        if missing_required:
            errors.append(f"Missing required column(s): {', '.join(missing_required)}")
        else:
            valid_rows, errors = _validate_rows(entity, mapping, rows)
        sheets.append(IngestSheetPreview(
            sheet=sheet_name, entity=entity, rows_total=len(rows), rows_valid=len(valid_rows),
            mapped={canon: headers[idx] for canon, idx in mapping.items()},
            unmapped=unmapped, errors=errors[:25], sample=valid_rows[:5],
        ))
        staged_sheets.append({"sheet": sheet_name, "entity": entity, "rows": valid_rows})

    staging_id = uuid.uuid4().hex[:12]
    _STAGING[staging_id] = {"at": time.monotonic(), "tenant_id": user.tenant_id, "sheets": staged_sheets}
    return IngestPreviewReply(
        staging_id=staging_id,
        filename=filename,
        sheets=sheets,
        total_valid=sum(s.rows_valid for s in sheets),
        total_rows=sum(s.rows_total for s in sheets),
    )


def commit(staging_id: str, user: User, default_project_id: Optional[str] = None) -> IngestCommitReply:
    from ._cache import invalidate_all
    from .audit import emit
    from .planning import get_project, upsert_bom_item, upsert_project
    from .vendor_store import add_supplier

    staged = _STAGING.get(staging_id)
    if staged is None:
        raise KeyError("staging_id not found or expired — re-run preview")
    if staged["tenant_id"] != user.tenant_id:
        raise KeyError("staging_id not found or expired — re-run preview")  # cross-tenant looks identical

    created = {"projects": 0, "bom_items": 0, "suppliers": 0}
    errors: List[str] = []
    refs: List[str] = []
    tenant = user.tenant_id

    # Pass 1: projects (BOM rows may reference them)
    for sheet in staged["sheets"]:
        if sheet["entity"] != "project":
            continue
        for row in sheet["rows"]:
            existing = get_project(row["project_id"])  # any tenant
            if existing is not None and existing.tenant_id != tenant:
                errors.append(
                    f"Project '{row['project_id']}' belongs to another tenant — skipped"
                )
                continue
            project = Project(
                project_id=row["project_id"], tenant_id=tenant, name=row["name"],
                client=row["client"], site=row["site"], sector=row["sector"],
                start_date=row["start_date"],
                milestones=[Milestone(code="M1", name="Engineering freeze", phase="engineering",
                                      required_on_site_date=row["start_date"])],
            )
            upsert_project(project)
            created["projects"] += 1
            refs.append(row["project_id"])

    # Pass 2: BOM lines
    seq = 0
    for sheet in staged["sheets"]:
        if sheet["entity"] != "bom":
            continue
        for row in sheet["rows"]:
            pid = row.get("project_id") or default_project_id
            if not pid:
                errors.append(f"BOM '{row['code']}': no project_id column and no default project selected — skipped")
                continue
            project = get_project(pid, tenant_id=tenant)
            if project is None:
                errors.append(f"BOM '{row['code']}': project {pid} not found in your tenant — skipped")
                continue
            seq += 1
            item_id = row.get("bom_item_id") or f"{pid}-ING{seq:04d}"
            upsert_bom_item(BOMItem(
                bom_item_id=item_id, tenant_id=tenant, project_id=pid,
                parent_item_id=row.get("parent_item_id"),
                code=row["code"], description=row["description"],
                quantity=row["quantity"], uom=row["uom"],
                unit_cost_usd=row.get("unit_cost_usd"),
                supplier_name=row.get("supplier_name"),
                spec_doc_id=row.get("spec_doc_id"), drawing_id=row.get("drawing_id"),
                long_lead_days=row.get("long_lead_days"),
                planned_need_date=row.get("planned_need_date"),
                milestone_code=row.get("milestone_code"),
                status=row["status"],
            ))
            created["bom_items"] += 1

    # Pass 3: suppliers
    for sheet in staged["sheets"]:
        if sheet["entity"] != "supplier":
            continue
        for row in sheet["rows"]:
            add_supplier(tenant, SupplierRecord(**row))
            created["suppliers"] += 1
            refs.append(row["name"])

    del _STAGING[staging_id]
    invalidate_all()
    emit(
        action="uploaded",
        entity_kind="project",
        entity_id=staging_id,
        subject="bulk ingest",
        summary=(
            f"Ingest committed: {created['projects']} project(s), "
            f"{created['bom_items']} BOM line(s), {created['suppliers']} supplier(s)"
        ),
        actor=user.user_id,
        source="csv_upload",
        tenant_id=tenant,
        metadata={**created, "errors": len(errors)},
    )
    return IngestCommitReply(created=created, errors=errors, refs=refs[:20])
