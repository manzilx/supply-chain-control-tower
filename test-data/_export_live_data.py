"""Pull the live seeded data from the running API into one Excel workbook.

Hits https://localhost (self-signed cert OK) as each tenant admin and
collects projects, BOM, vendors, sourcing, commercial, approvals, alerts.
Writes test-data/control-tower-synthetic-data.xlsx.
"""

from __future__ import annotations

import json
import ssl
import urllib.request
import urllib.error
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = "https://localhost"
CTX = ssl.create_default_context()
CTX.check_hostname = False
CTX.verify_mode = ssl.CERT_NONE

TENANTS = [
    ("arcforge", "arcforge-admin-01"),
    ("helios", "helios-admin-01"),
    ("northwind", "northwind-admin-01"),
]

HERE = Path(__file__).parent


def _login(user_id: str) -> str:
    req = urllib.request.Request(
        f"{BASE}/api/auth/login", method="POST",
        data=json.dumps({"user_id": user_id}).encode(),
    )
    req.add_header("Content-Type", "application/json")
    return json.loads(urllib.request.urlopen(req, context=CTX).read())["token"]


def _get(path: str, token: str):
    req = urllib.request.Request(f"{BASE}{path}")
    req.add_header("Authorization", f"Bearer {token}")
    try:
        return json.loads(urllib.request.urlopen(req, context=CTX).read())
    except urllib.error.HTTPError:
        return None


# ----- styling ---------------------------------------------------------------

FONT = "Arial"
NAVY = "1B2D44"
ACCENT = "11B29E"
LINE = "D9DDE2"
ZEBRA = "F4F7FA"

HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
BODY_FONT = Font(name=FONT, size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=15)
EYEBROW_FONT = Font(name=FONT, bold=True, color=ACCENT, size=9)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(border_style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Per-tenant row tint so you can eyeball isolation at a glance.
TENANT_FILL = {
    "arcforge": PatternFill("solid", fgColor="EAF6F3"),
    "helios": PatternFill("solid", fgColor="EAF0F7"),
    "northwind": PatternFill("solid", fgColor="F6F1EA"),
}


def write_sheet(ws, headers, rows, title, eyebrow, widths, tenant_col=0):
    r = 1
    ws.cell(row=r, column=1, value=eyebrow).font = EYEBROW_FONT
    r += 1
    ws.cell(row=r, column=1, value=title).font = TITLE_FONT
    r += 2
    header_row = r
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[r].height = 26
    for row in rows:
        r += 1
        tenant = str(row[tenant_col]).lower() if tenant_col is not None and row else ""
        tint = TENANT_FILL.get(tenant)
        for ci, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font = BODY_FONT
            c.alignment = WRAP
            c.border = BOX
            if tint:
                c.fill = tint
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    return r


def main() -> None:
    wb = Workbook()

    # collect everything keyed by tenant
    data = {}
    for tid, admin in TENANTS:
        tok = _login(admin)
        d = {"token": tok}
        d["projects"] = _get("/api/projects", tok) or []
        d["progress"] = {p["project_id"]: p for p in (_get("/api/projects/progress", tok) or [])}
        d["bom"] = {
            p["project_id"]: (_get(f"/api/projects/{p['project_id']}/bom", tok) or [])
            for p in d["projects"]
        }
        d["vendors"] = _get("/api/vendors/intel", tok) or []
        d["concentration"] = _get("/api/vendors/concentration", tok) or []
        d["prs"] = _get("/api/prs", tok) or []
        d["rfqs"] = _get("/api/rfqs", tok) or []
        d["pos"] = _get("/api/sourcing-pos", tok) or []
        d["commercial"] = _get("/api/commercial/summary", tok) or {}
        d["approvals"] = _get("/api/approvals", tok) or []
        d["alerts"] = (_get("/api/alerts", tok) or {}).get("alerts", [])
        d["demo"] = _get("/api/demo", tok) or {}
        data[tid] = d

    # ---- README ----
    ws = wb.active
    ws.title = "README"
    ws.cell(row=1, column=1, value="SUPPLY CHAIN CONTROL TOWER").font = EYEBROW_FONT
    ws.cell(row=2, column=1, value="Synthetic Data — live export").font = TITLE_FONT
    notes = [
        ("Source", "Pulled live from the running API across all 3 tenants. Reflects the seed code in app/sample_data.py + app/planning.py + fixtures/hydro/."),
        ("Tenants", "arcforge (Power Systems EPC) · helios (Offshore Engineering) · northwind (Heavy Engineering). Rows are tinted by tenant."),
        ("Isolation", "Each tenant's data is fully scoped — these rows are what that tenant actually sees via its own token."),
        ("Sheets", "Projects · BOM (every line, all projects) · Vendors · Concentration · PRs · POs · Commercial · Approvals · Alerts."),
        ("Seeded counts", " · ".join(
            f"{tid}: {len(d['projects'])} projects / "
            f"{sum(len(b) for b in d['bom'].values())} BOM lines / "
            f"{len(d['vendors'])} vendors"
            for tid, d in data.items()
        )),
    ]
    r = 4
    for k, v in notes:
        ws.cell(row=r, column=1, value=k).font = Font(name=FONT, bold=True, size=11)
        c = ws.cell(row=r, column=2, value=v)
        c.font = BODY_FONT
        c.alignment = WRAP
        ws.row_dimensions[r].height = max(30, 15 * (1 + len(v) // 90))
        r += 1
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 120

    # ---- Projects ----
    ws = wb.create_sheet("Projects")
    rows = []
    for tid, d in data.items():
        for p in d["projects"]:
            prog = d["progress"].get(p["project_id"], {})
            rows.append([
                tid, p["project_id"], p["name"], p.get("client"), p.get("site"),
                p.get("sector"), p.get("start_date"), len(p.get("milestones", [])),
                len(d["bom"].get(p["project_id"], [])),
                prog.get("completion_pct"), prog.get("milestones_passed"),
                prog.get("bom_delivered"), round(prog.get("budget_value_usd", 0)),
            ])
    write_sheet(
        ws,
        ["Tenant", "Project ID", "Name", "Client", "Site", "Sector", "Start",
         "Milestones", "BOM lines", "Completion %", "MS passed", "BOM delivered", "BOM budget $"],
        rows, "Projects", "PORTFOLIO",
        [12, 18, 42, 30, 28, 24, 12, 11, 10, 13, 11, 14, 16],
    )

    # ---- BOM (every line) ----
    ws = wb.create_sheet("BOM")
    rows = []
    for tid, d in data.items():
        for pid, items in d["bom"].items():
            for b in items:
                rows.append([
                    tid, pid, b.get("bom_item_id"), b.get("code"), b.get("description"),
                    b.get("category"), b.get("quantity"), b.get("uom"), b.get("unit_cost_usd"),
                    b.get("supplier_name"), b.get("long_lead_days"), b.get("planned_need_date"),
                    b.get("milestone_code"), b.get("status"),
                ])
    write_sheet(
        ws,
        ["Tenant", "Project", "Item ID", "Code", "Description", "Category", "Qty", "UoM",
         "Unit cost $", "Supplier", "Lead (d)", "Need by", "Milestone", "Status"],
        rows, f"Bill of Materials — {len(rows)} lines", "BOM",
        [12, 16, 16, 18, 36, 20, 7, 7, 13, 24, 9, 13, 11, 13],
    )

    # ---- Vendors ----
    ws = wb.create_sheet("Vendors")
    rows = []
    for tid, d in data.items():
        for v in d["vendors"]:
            rows.append([
                tid, v.get("vendor"), v.get("category"), v.get("country"),
                v.get("composite_score"), v.get("composite_grade"),
                v.get("on_time_delivery_pct"), v.get("quality_ppm"),
                round(v.get("annual_spend_usd", 0)), v.get("flags_count"),
                "yes" if v.get("single_source_exposure") else "",
            ])
    write_sheet(
        ws,
        ["Tenant", "Vendor", "Category", "Country", "Score", "Grade",
         "OTD %", "Quality PPM", "Annual spend $", "Flags", "Single-source"],
        rows, "Vendor Scorecards", "VENDORS",
        [12, 30, 24, 16, 9, 8, 9, 12, 16, 8, 14],
    )

    # ---- Concentration ----
    ws = wb.create_sheet("Concentration")
    rows = []
    for tid, d in data.items():
        for c in d["concentration"]:
            rows.append([
                tid, c.get("category"), c.get("vendor_count"),
                round(c.get("total_spend_usd", 0)), c.get("top_vendor"),
                c.get("top_vendor_share_pct"), "yes" if c.get("single_source") else "",
            ])
    write_sheet(
        ws,
        ["Tenant", "Category", "# Vendors", "Total spend $", "Top vendor", "Top share %", "Single-source"],
        rows, "Category Concentration", "VENDORS",
        [12, 26, 11, 16, 30, 13, 14],
    )

    # ---- PRs ----
    ws = wb.create_sheet("PRs")
    rows = []
    for tid, d in data.items():
        for p in d["prs"]:
            rows.append([
                tid, p.get("pr_no"), p.get("project_id"), p.get("code"),
                p.get("quantity"), p.get("uom"), round(p.get("budget_value_usd") or 0),
                p.get("buyer"), p.get("strategy"), p.get("status"),
                p.get("rfq_no"), p.get("po_no"),
            ])
    write_sheet(
        ws,
        ["Tenant", "PR No", "Project", "Code", "Qty", "UoM", "Budget $", "Buyer",
         "Strategy", "Status", "RFQ", "PO"],
        rows, "Purchase Requisitions", "SOURCING",
        [12, 12, 16, 18, 8, 7, 13, 16, 14, 12, 12, 12],
    )

    # ---- POs ----
    ws = wb.create_sheet("POs")
    rows = []
    for tid, d in data.items():
        for p in d["pos"]:
            rows.append([
                tid, p.get("po_no"), p.get("project_id"), p.get("code"),
                p.get("vendor"), p.get("quantity"), round(p.get("value_usd") or 0),
                p.get("incoterm"), p.get("lead_time_days"), p.get("status"),
            ])
    write_sheet(
        ws,
        ["Tenant", "PO No", "Project", "Code", "Vendor", "Qty", "Value $",
         "Incoterm", "Lead (d)", "Status"],
        rows, "Sourcing Purchase Orders", "SOURCING",
        [12, 14, 16, 18, 26, 8, 14, 10, 9, 12],
    )

    # ---- Commercial ----
    ws = wb.create_sheet("Commercial")
    rows = []
    for tid, d in data.items():
        c = d["commercial"]
        for proj in c.get("projects", []):
            rows.append([
                tid, proj.get("project_name"), proj.get("line_count"),
                round(proj.get("total_budget_usd", 0)), round(proj.get("total_quoted_usd", 0)),
                round(proj.get("total_awarded_usd", 0)), round(proj.get("total_savings_usd", 0)),
                proj.get("variance_pct"), proj.get("over_budget_lines"),
            ])
    write_sheet(
        ws,
        ["Tenant", "Project", "Lines", "Budget $", "Quoted $", "Awarded $",
         "Savings $", "Variance %", "Over-budget lines"],
        rows, "Commercial Roll-up", "COMMERCIAL",
        [12, 42, 8, 15, 15, 15, 14, 12, 16],
    )

    # ---- Approvals ----
    ws = wb.create_sheet("Approvals")
    rows = []
    for tid, d in data.items():
        for a in d["approvals"]:
            rows.append([
                tid, a.get("approval_id"), a.get("kind"), a.get("title"),
                a.get("status"), a.get("requested_by_name"), a.get("decided_by_name"),
                a.get("result_ref"),
            ])
    if not rows:
        rows = [["—", "—", "—", "No approvals raised yet (award a single-source RFQ to create one)", "", "", "", ""]]
    write_sheet(
        ws,
        ["Tenant", "ID", "Kind", "Title", "Status", "Requested by", "Decided by", "Result ref"],
        rows, "Approvals Queue", "GOVERNANCE",
        [12, 10, 20, 40, 14, 22, 22, 14],
    )

    # ---- Alerts ----
    ws = wb.create_sheet("Alerts")
    rows = []
    for tid, d in data.items():
        for a in d["alerts"]:
            rows.append([
                tid, a.get("severity"), a.get("category"), a.get("title"), a.get("detail"),
            ])
    write_sheet(
        ws,
        ["Tenant", "Severity", "Category", "Title", "Detail"],
        rows, "Live Alert Feed", "MONITOR",
        [12, 11, 13, 44, 60],
    )

    out = HERE / "control-tower-synthetic-data.xlsx"
    wb.save(out)
    print(f"Wrote {out.name} ({out.stat().st_size:,} bytes)")
    for tid, d in data.items():
        print(f"  {tid}: {len(d['projects'])} projects, "
              f"{sum(len(b) for b in d['bom'].values())} BOM, "
              f"{len(d['vendors'])} vendors, {len(d['prs'])} PRs, "
              f"{len(d['pos'])} POs, {len(d['approvals'])} approvals, {len(d['alerts'])} alerts")


if __name__ == "__main__":
    main()
