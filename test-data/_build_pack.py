"""Build the comprehensive test pack — Excel workbook + uploadable BOM CSVs.

Outputs (relative to test-data/):
  control-tower-test-pack.xlsx   (multi-sheet workbook)
  bom-arcforge-PRJ-RB-660.csv    (uploadable to Riverbank)
  bom-helios-PRJ-NS-OSS.csv      (uploadable to North Sea Substation)
  bom-northwind-HYD-MAHADEV-220.csv (uploadable to Mahadev Hydro)
"""

from __future__ import annotations

import csv
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = Path(__file__).parent

# ----- styling ---------------------------------------------------------------

FONT = "Arial"
NAVY = "1B2D44"
ACCENT = "11B29E"
LINE = "D9DDE2"
ZEBRA = "F4F7FA"

HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
BODY_FONT = Font(name=FONT, size=10)
EYEBROW_FONT = Font(name=FONT, bold=True, color=ACCENT, size=9)
TITLE_FONT = Font(name=FONT, bold=True, size=16)
SECTION_FONT = Font(name=FONT, bold=True, size=12)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")

THIN = Side(border_style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row: int, last_col: int):
    for col in range(1, last_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[row].height = 22


def style_body(ws, first_row: int, last_row: int, last_col: int, zebra: bool = True):
    for r in range(first_row, last_row + 1):
        for col in range(1, last_col + 1):
            c = ws.cell(row=r, column=col)
            c.font = BODY_FONT
            c.alignment = WRAP
            c.border = BOX
            if zebra and (r - first_row) % 2 == 1:
                c.fill = PatternFill("solid", fgColor=ZEBRA)


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_table(ws, *, headers, rows, title=None, eyebrow=None, widths=None, start_row=1):
    r = start_row
    if eyebrow:
        ws.cell(row=r, column=1, value=eyebrow).font = EYEBROW_FONT
        r += 1
    if title:
        ws.cell(row=r, column=1, value=title).font = TITLE_FONT
        r += 2
    header_row = r
    for ci, h in enumerate(headers, start=1):
        ws.cell(row=r, column=ci, value=h)
    style_header(ws, r, len(headers))
    for row in rows:
        r += 1
        for ci, v in enumerate(row, start=1):
            ws.cell(row=r, column=ci, value=v)
    style_body(ws, header_row + 1, r, len(headers))
    if widths:
        set_widths(ws, widths)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    return r


# ----- data ------------------------------------------------------------------

TODAY = date.today()


def d(days: int) -> str:
    return (TODAY + timedelta(days=days)).isoformat()


TENANTS = [
    ("arcforge", "Arcforge Engineering", "Power Systems EPC"),
    ("helios", "Helios Offshore", "Offshore Engineering"),
    ("northwind", "Northwind Heavy Engineering", "Heavy Engineering"),
]

ROLES = [
    ("admin", "Admin", "All read + write across modules + tenant switch (admin only)."),
    ("procurement_head", "Procurement Head", "Read all + manage PR/RFQ/Award/PO + decide approvals."),
    ("buyer", "Buyer", "Read all + create PR, RFQ, Quote. Cannot award POs over threshold."),
    ("expeditor", "Lead Expeditor", "Read all + create follow-up emails + shipment events."),
    ("viewer", "Project Controls Viewer", "Read-only access to every page."),
]


def personas():
    rows = []
    for tid, tname, _ in TENANTS:
        for i, (role, label, what) in enumerate(ROLES, start=1):
            user_id = f"{tid}-{role.split('_')[0]}-0{i}".replace("admin-02", "admin-01")
            # Re-derive expected ids deterministically: tenant-roleshort-01
            short = {"admin": "admin", "procurement_head": "head", "buyer": "buyer",
                     "expeditor": "expeditor", "viewer": "viewer"}[role]
            user_id = f"{tid}-{short}-01"
            email = f"{short}@{('helios-offshore' if tid=='helios' else 'northwind' if tid=='northwind' else 'arcforge')}.co{'m' if tid!='northwind' else ''}"
            rows.append([user_id, label, email, tname, role, what])
    return rows


PROJECTS = [
    # tenant, project_id, name, client, site, sector, expected_completion, expected_bom_lines
    ("arcforge", "PRJ-RB-660", "Riverbank 2x660 MW Power Plant", "Arcforge Power", "Riverbank, Gujarat", "Power EPC", "0%", 8),
    ("arcforge", "PRJ-AF-CCGT", "Tanjore Combined-Cycle Gas Plant 750 MW", "Arcforge Power", "Tanjore, Tamil Nadu", "Power EPC", "0%", 5),
    ("arcforge", "PRJ-AF-SUB", "Sundarpur 765 kV Substation Upgrade", "Arcforge Grid", "Sundarpur, Madhya Pradesh", "Transmission EPC", "0%", 5),
    ("arcforge", "PRJ-AF-MERIDIAN", "Meridian 2x350 MW CCGT — Unit 1 Execution", "Arcforge Power", "Meridian, Andhra Pradesh", "Power EPC", "~56% (in-flight)", 9),
    ("helios", "PRJ-NS-OSS", "North Sea Offshore Substation", "Helios Offshore Operations", "Aberdeen, UK", "Offshore Engineering", "0%", 4),
    ("helios", "PRJ-HE-WIND", "Dogger Bank Offshore Wind 480 MW", "Helios Renewables", "Dogger Bank, North Sea", "Offshore Wind EPC", "0%", 5),
    ("helios", "PRJ-HE-FPSO", "Hawthorn FPSO Topsides", "Helios Upstream", "West of Shetland", "Offshore Oil & Gas EPC", "0%", 5),
    ("helios", "PRJ-HE-VALHALL", "Valhall Bravo Platform Tie-In", "Helios Upstream", "Valhall Field, North Sea", "Offshore Oil & Gas EPC", "~56% (in-flight)", 9),
    ("northwind", "HYD-MAHADEV-220", "Mahadev Hydro 220 MW (2x110)", "Northern Hills Power Corporation", "Sutlej Basin, Himachal Pradesh", "Hydropower EPC", "~17%", 70),
    ("northwind", "PRJ-NW-STEEL", "Polaris Steel Rolling Mill Modernization", "Polaris Steel", "Jamshedpur, Jharkhand", "Heavy Industry EPC", "0%", 5),
    ("northwind", "PRJ-NW-CEMENT", "Granite Ridge Cement Line 2", "Granite Ridge Cement", "Chandrapur, Maharashtra", "Heavy Industry EPC", "0%", 5),
    ("northwind", "PRJ-NW-KAVI", "Kavi Hydro Unit 3 Major Overhaul", "Northern Hills Power Corporation", "Kavi, Himachal Pradesh", "Hydropower EPC", "~53% (in-flight)", 9),
]


# Comprehensive BOM library — uploadable per project.
# Columns: code, description, category, quantity, uom, unit_cost_usd,
# supplier_name, spec_doc_id, drawing_id, long_lead_days, planned_need_date,
# milestone_code  (status defaults to spec_missing when spec_doc_id is empty)

BOM_ARCFORGE = [  # for PRJ-RB-660 (Riverbank)
    ("BFP-660-B", "Boiler feed pump set B", "Rotating equipment", 2, "EA", 920000, "Helios Cast & Forge", "SPEC-RB-001", "DRG-RB-001", 240, d(180), "M3"),
    ("DEAERATOR-660", "Deaerator + storage tank", "Process", 1, "EA", 380000, "Kerala Forge Works", "SPEC-RB-001", "", 180, d(150), "M3"),
    ("CONDENSER-AC", "Air-cooled condenser bank", "Process", 8, "EA", 220000, None, "", "", 270, d(200), "M4"),
    ("HRSG-MODULE", "HRSG pressure module", "Heat recovery", 4, "EA", 1100000, "Helios Cast & Forge", "SPEC-RB-001", "", 300, d(210), "M3"),
    ("ECONOMIZER-COIL", "Economizer coil bundle", "Heat recovery", 4, "EA", 180000, "Helios Cast & Forge", "", "", 210, d(150), "M3"),
    ("DESH-PIPING-CR", "Desuperheater piping chrome-moly", "Piping", 320, "M", 1200, "Copperline Metals", "", "", 90, d(120), "M3"),
    ("VALVE-24-A105", "24-inch forged gate valve", "Forged valves", 12, "EA", 6800, "Kerala Forge Works", "SPEC-RB-002", "", 120, d(100), "M3"),
    ("RELIEF-VALVE", "Steam relief valve assembly", "Forged valves", 24, "EA", 2400, "Kerala Forge Works", "", "", 90, d(110), "M3"),
    ("ACTUATOR-PNEU", "Pneumatic valve actuator", "Automation", 36, "EA", 1800, "Mitsuba Automation", "", "", 75, d(85), "M2"),
    ("SOOT-BLOWER", "Wall-tube soot blower set", "Process", 16, "EA", 14000, None, "", "", 150, d(180), "M3"),
    ("HV-SWITCHGEAR-11", "11 kV switchgear panel", "Electrical", 6, "EA", 95000, "BluePeak Controls", "SPEC-RB-001", "", 180, d(160), "M3"),
    ("CT-MV-630", "MV current transformer 630A", "Electrical", 24, "EA", 4200, "BluePeak Controls", "", "", 90, d(110), "M3"),
    ("MOTOR-IM-3300", "3.3 kV induction motor 4 MW", "Rotating equipment", 6, "EA", 165000, "Helios Cast & Forge", "", "", 180, d(150), "M3"),
    ("PLC-S7-CPU", "PLC CPU module 1500 series", "Automation", 8, "EA", 8400, "BluePeak Controls", "", "", 60, d(80), "M2"),
    ("CABLE-MV-XLPE", "33 kV XLPE single-core cable", "Cables", 18, "KM", 92000, "Copperline Metals", "", "", 75, d(110), "M3"),
    ("CABLE-CONTROL", "Multipair instrumentation cable", "Cables", 42, "KM", 6800, "Copperline Metals", "", "", 60, d(100), "M3"),
    ("CONDUIT-GAL", "Hot-dip galvanized conduit", "Electrical", 8400, "M", 22, "Delhi Metals Co", "", "", 30, d(60), "M2"),
    ("RACEWAY-CABLE", "Cable raceway galvanized", "Electrical", 2200, "M", 86, "Delhi Metals Co", "", "", 30, d(60), "M2"),
    ("INSTR-PRES", "Pressure transmitter HART", "Instrumentation", 240, "EA", 1100, "Mitsuba Automation", "", "", 75, d(80), "M2"),
    ("INSTR-TEMP-RTD", "RTD temperature sensor", "Instrumentation", 380, "EA", 240, "Mitsuba Automation", "", "", 60, d(75), "M2"),
    ("INSTR-FLOW-VTX", "Vortex flow meter", "Instrumentation", 64, "EA", 4800, "Mitsuba Automation", "", "", 75, d(85), "M2"),
    ("BALL-VALVE-2", '2"-150 ball valve carbon', "Forged valves", 480, "EA", 320, "Kerala Forge Works", "", "", 60, d(90), "M2"),
    ("PSV-150-HF", "Pilot-operated PSV", "Forged valves", 36, "EA", 4800, None, "", "", 120, d(120), "M3"),
    ("UPS-CRITICAL", "100 kVA critical load UPS", "Electrical", 4, "EA", 78000, "BluePeak Controls", "", "", 90, d(110), "M3"),
    ("BATTERY-VRLA", "VRLA battery bank 220V", "Electrical", 6, "EA", 14000, "BluePeak Controls", "", "", 60, d(90), "M3"),
    ("FIRE-DELUGE", "Deluge valve fire-water", "Fire", 18, "EA", 6200, None, "", "", 90, d(120), "M3"),
    ("FIRE-CO2-CYL", "CO2 cylinder bank", "Fire", 24, "EA", 980, None, "", "", 45, d(90), "M3"),
    ("PIPING-CS-A106", "A106-B carbon steel pipe", "Piping", 1800, "M", 320, "Copperline Metals", "", "", 75, d(110), "M3"),
    ("PIPING-SS-316", "SS316L stainless pipe", "Piping", 420, "M", 980, "Copperline Metals", "", "", 90, d(110), "M3"),
    ("CRANE-EOT-50", "EOT crane 50T", "Crane / lifting", 2, "EA", 285000, None, "", "", 240, d(200), "M4"),
    ("HVAC-AHU-30K", "AHU 30,000 CFM", "HVAC", 6, "EA", 64000, None, "", "", 120, d(150), "M4"),
    ("CW-PUMP-VS", "Cooling water vertical pump", "Cooling", 4, "EA", 240000, "Helios Cast & Forge", "", "", 180, d(150), "M3"),
    ("OIL-LUBE-SKID", "Lube oil skid 200 GPM", "Lubrication", 2, "EA", 96000, None, "", "", 150, d(160), "M3"),
    ("STRUCT-BEAM", "Structural beam W-section", "Structural", 580, "T", 1850, None, "", "", 60, d(90), "M3"),
    ("REBAR-BUNDLE", "Rebar #8 bundle", "Civil", 1200, "T", 720, None, "", "", 30, d(60), "M2"),
    ("EARTH-CABLE-CU", "Copper earthing cable", "Electrical", 6800, "M", 28, "Copperline Metals", "", "", 45, d(90), "M2"),
    ("LIGHTING-FIX-LED", "LED industrial light fixture", "Electrical", 480, "EA", 280, "Delhi Metals Co", "", "", 30, d(60), "M2"),
    ("PORTABLE-WELDER", "300A portable welding set", "Tools", 14, "EA", 4800, None, "", "", 60, d(80), "M2"),
]

BOM_HELIOS = [  # for PRJ-NS-OSS (North Sea Substation)
    ("GIS-66-BAY", "GIS 66 kV bay extension", "Electrical", 4, "EA", 880000, "BluePeak Controls", "SPEC-NS-001", "DRG-NS-001", 280, d(180), "M3"),
    ("XFMR-PWR-220", "Power transformer 220 MVA", "Electrical", 2, "EA", 2400000, None, "", "", 380, d(200), "M3"),
    ("REACTOR-SH", "Shunt reactor 200 MVAr", "Electrical", 2, "EA", 920000, "BluePeak Controls", "", "", 300, d(200), "M3"),
    ("UMBILICAL-DYN", "Dynamic umbilical cable", "Subsea cables", 12, "KM", 480000, "NorthCable Subsea", "SPEC-NS-001", "", 360, d(220), "M3"),
    ("J-TUBE-CRA", "J-tube CRA lined", "Tubing", 12, "EA", 95000, "Copperline Metals", "", "", 90, d(140), "M3"),
    ("BUSHING-66KV", "66 kV transformer bushing", "Electrical", 12, "EA", 38000, "BluePeak Controls", "", "", 180, d(160), "M3"),
    ("SURGE-ARRESTER", "Surge arrester 66 kV", "Electrical", 18, "EA", 8400, "BluePeak Controls", "", "", 90, d(120), "M3"),
    ("CABLE-LV-SHIP", "LV shipboard armoured cable", "Cables", 32, "KM", 64000, "NorthCable Subsea", "", "", 75, d(120), "M3"),
    ("DECK-PLATING", "Deck plating EH36 grade", "Offshore structural", 380, "T", 2900, "Aberdeen Steel Fab", "", "", 120, d(150), "M2"),
    ("PILE-SLEEVE", "Pile sleeve fabrication", "Offshore structural", 8, "EA", 240000, "Aberdeen Steel Fab", "", "", 180, d(200), "M3"),
    ("PADEYE-MAIN", "Main lifting padeye", "Offshore structural", 16, "EA", 18000, "Aberdeen Steel Fab", "", "", 90, d(120), "M2"),
    ("COATING-IZ", "Inorganic zinc primer", "Marine epoxy", 1800, "L", 22, "Coastline Marine Coatings", "", "", 45, d(75), "M2"),
    ("COATING-EPOXY", "Epoxy mid coat", "Marine epoxy", 1200, "L", 28, "Coastline Marine Coatings", "", "", 45, d(75), "M2"),
    ("ROV-PANEL", "ROV interface panel", "ROV connectors", 6, "EA", 64000, "DeepDive ROV Systems", "", "", 120, d(150), "M3"),
    ("BUOY-SUB", "Subsea distributed buoyancy", "ROV connectors", 24, "EA", 8400, "DeepDive ROV Systems", "", "", 90, d(120), "M3"),
    ("FENDER-DOLPHIN", "Vessel fender + dolphin", "Marine equipment", 8, "EA", 96000, "Aberdeen Steel Fab", "", "", 120, d(150), "M3"),
    ("WALKWAY-GRATING", "GRP walkway grating", "Offshore structural", 1800, "M2", 280, "Aberdeen Steel Fab", "", "", 60, d(100), "M2"),
    ("HANDRAIL-SS", "SS316 handrail tubular", "Offshore structural", 2800, "M", 86, "Copperline Metals", "", "", 45, d(90), "M2"),
    ("LIFEBOAT-DAVIT", "Lifeboat davit launching system", "Marine equipment", 4, "EA", 320000, None, "", "", 240, d(220), "M3"),
    ("LIFERAFT-25P", "25-person life raft container", "Marine equipment", 12, "EA", 18000, None, "", "", 60, d(120), "M3"),
    ("VENT-AHU-EX", "Ex-rated AHU", "HVAC", 6, "EA", 84000, None, "", "", 120, d(150), "M3"),
    ("HVAC-DAMPER-FG", "Fire/gas damper", "HVAC", 28, "EA", 4200, None, "", "", 90, d(120), "M3"),
    ("FW-PUMP-ELEC", "Electric firewater pump", "Fire", 4, "EA", 240000, None, "", "", 180, d(160), "M3"),
    ("GAS-DETECT-IR", "IR gas detector", "Instrumentation", 96, "EA", 2800, "DeepDive ROV Systems", "", "", 90, d(110), "M3"),
    ("F&G-PANEL", "F&G control panel", "Automation", 4, "EA", 145000, "BluePeak Controls", "", "", 150, d(150), "M3"),
    ("PAGA-SYSTEM", "PAGA public address", "Automation", 1, "LOT", 380000, "BluePeak Controls", "", "", 180, d(170), "M3"),
    ("HELI-DECK-NET", "Helideck perimeter safety net", "Offshore structural", 1, "LOT", 95000, "Aberdeen Steel Fab", "", "", 90, d(140), "M3"),
    ("MOORING-CHAIN", "R4 mooring chain 95mm", "Mooring", 1800, "M", 980, None, "", "", 270, d(200), "M3"),
    ("ANCHOR-DEA", "Drag embedment anchor", "Mooring", 8, "EA", 220000, None, "", "", 240, d(200), "M3"),
    ("DP-THRUSTER", "Azimuth thruster 3 MW", "Rotating equipment", 4, "EA", 1850000, None, "", "", 360, d(220), "M3"),
    ("CRANE-OFFSHORE", "Knuckle-boom crane 50T", "Crane / lifting", 2, "EA", 580000, None, "", "", 270, d(200), "M3"),
    ("CONN-WET-MATE", "ROV wet-mate 24-pin connector", "ROV connectors", 24, "EA", 5600, "DeepDive ROV Systems", "", "", 90, d(110), "M3"),
]

BOM_NORTHWIND = [  # for HYD-MAHADEV-220 (Mahadev Hydro)
    ("RUNNER-FRANCIS-A", "Francis turbine runner Unit A", "Hydro turbines", 1, "EA", 2400000, "Andritz Hydro", "SPEC-HYD-001", "DRG-HYD-RUN", 420, d(280), "M07"),
    ("RUNNER-FRANCIS-B", "Francis turbine runner Unit B", "Hydro turbines", 1, "EA", 2400000, "Andritz Hydro", "SPEC-HYD-001", "DRG-HYD-RUN", 420, d(310), "M08"),
    ("WICKET-GATE-A", "Wicket gate assembly Unit A", "Hydro turbines", 24, "EA", 48000, "Andritz Hydro", "", "", 240, d(220), "M07"),
    ("WICKET-GATE-B", "Wicket gate assembly Unit B", "Hydro turbines", 24, "EA", 48000, "Andritz Hydro", "", "", 240, d(250), "M07"),
    ("SPIRAL-CASE-A", "Spiral casing Unit A", "Hydro turbines", 1, "EA", 980000, "Andritz Hydro", "", "", 300, d(220), "M07"),
    ("SPIRAL-CASE-B", "Spiral casing Unit B", "Hydro turbines", 1, "EA", 980000, "Andritz Hydro", "", "", 300, d(250), "M07"),
    ("DRAFT-TUBE-A", "Draft tube fabrication Unit A", "Hydro turbines", 1, "EA", 320000, "Voith Industrial Services", "", "", 240, d(220), "M07"),
    ("DRAFT-TUBE-B", "Draft tube fabrication Unit B", "Hydro turbines", 1, "EA", 320000, "Voith Industrial Services", "", "", 240, d(250), "M07"),
    ("GEN-STATOR-A", "Generator stator assembly A", "Generators", 1, "EA", 1850000, "Bharat Heavy Electricals", "SPEC-HYD-002", "", 420, d(290), "M08"),
    ("GEN-STATOR-B", "Generator stator assembly B", "Generators", 1, "EA", 1850000, "Bharat Heavy Electricals", "SPEC-HYD-002", "", 420, d(320), "M08"),
    ("GEN-ROTOR-A", "Generator rotor + poles A", "Generators", 1, "EA", 1450000, "Bharat Heavy Electricals", "", "", 420, d(290), "M08"),
    ("GEN-ROTOR-B", "Generator rotor + poles B", "Generators", 1, "EA", 1450000, "Bharat Heavy Electricals", "", "", 420, d(320), "M08"),
    ("EXCITER-A", "Brushless exciter Unit A", "Generators", 1, "EA", 240000, "Bharat Heavy Electricals", "", "", 240, d(280), "M08"),
    ("EXCITER-B", "Brushless exciter Unit B", "Generators", 1, "EA", 240000, "Bharat Heavy Electricals", "", "", 240, d(310), "M08"),
    ("XFMR-GSU-160", "Generator step-up transformer 160 MVA", "Electrical", 2, "EA", 1100000, None, "", "", 360, d(300), "M09"),
    ("XFMR-AUX", "Auxiliary transformer 5 MVA", "Electrical", 4, "EA", 145000, "Bharat Heavy Electricals", "", "", 180, d(180), "M07"),
    ("GIS-220-BAY", "220 kV GIS bay", "Electrical", 6, "EA", 1450000, None, "", "", 320, d(330), "M10"),
    ("GIS-66-FEEDER", "66 kV GIS feeder bay", "Electrical", 4, "EA", 380000, None, "", "", 240, d(300), "M09"),
    ("PROT-RELAY-A", "Generator protection relay panel", "Automation", 2, "EA", 96000, None, "", "", 180, d(300), "M09"),
    ("GOV-DIGITAL-A", "Digital governor + hydraulics A", "Automation", 1, "EA", 620000, "Voith Industrial Services", "", "", 240, d(280), "M07"),
    ("GOV-DIGITAL-B", "Digital governor + hydraulics B", "Automation", 1, "EA", 620000, "Voith Industrial Services", "", "", 240, d(310), "M07"),
    ("PENSTOCK-LINER-1", "Penstock liner segment 1", "Penstock fabrication", 6, "EA", 380000, "Voith Industrial Services", "SPEC-HYD-003", "", 300, d(150), "M06"),
    ("PENSTOCK-LINER-2", "Penstock liner segment 2", "Penstock fabrication", 6, "EA", 380000, "Voith Industrial Services", "", "", 300, d(170), "M06"),
    ("PENSTOCK-RING", "Penstock stiffening ring", "Penstock fabrication", 24, "EA", 28000, "Voith Industrial Services", "", "", 180, d(150), "M06"),
    ("EXP-JOINT", "Expansion joint bellows", "Piping", 8, "EA", 38000, None, "", "", 150, d(180), "M07"),
    ("BUTTERFLY-3M", "3 m butterfly intake valve", "Forged valves", 2, "EA", 980000, None, "", "", 300, d(220), "M07"),
    ("GATE-RADIAL", "Radial gate spillway", "Hydro turbines", 4, "EA", 480000, None, "", "", 270, d(170), "M06"),
    ("HOIST-RADIAL", "Spillway gate hoist", "Crane / lifting", 4, "EA", 220000, None, "", "", 240, d(180), "M06"),
    ("EOT-CRANE-200", "EOT crane 200T", "Crane / lifting", 1, "EA", 480000, None, "", "", 300, d(220), "M07"),
    ("EOT-CRANE-50-AUX", "EOT crane 50T auxiliary", "Crane / lifting", 2, "EA", 220000, None, "", "", 240, d(220), "M07"),
    ("COOLING-HEX", "Heat exchanger generator cooling", "Cooling", 4, "EA", 86000, None, "", "", 180, d(280), "M08"),
    ("OIL-PURIFIER", "Oil purification skid", "Lubrication", 2, "EA", 64000, None, "", "", 150, d(280), "M08"),
    ("FIRE-CO2-SYS", "CO2 fire suppression generator", "Fire", 4, "EA", 38000, None, "", "", 120, d(290), "M08"),
    ("HVAC-MACH-HALL", "Machine hall HVAC system", "HVAC", 1, "LOT", 580000, None, "", "", 180, d(300), "M09"),
    ("DCS-HYDRO", "DCS distributed control system", "Automation", 1, "LOT", 1850000, None, "", "", 240, d(310), "M09"),
    ("INSTR-VIB", "Vibration monitoring system", "Instrumentation", 12, "EA", 28000, None, "", "", 150, d(290), "M08"),
    ("INSTR-PT-DAM", "Dam piezometer / instrumentation", "Instrumentation", 240, "EA", 4800, None, "", "", 90, d(120), "M06"),
    ("STEEL-DAM-FACE", "Dam concrete face rebar", "Civil", 4800, "T", 720, None, "", "", 30, d(60), "M04"),
    ("CEMENT-MASS-CON", "Mass concrete cement", "Civil", 28000, "T", 92, None, "", "", 15, d(60), "M04"),
    ("AGGREGATE-CON", "Concrete aggregate graded", "Civil", 68000, "T", 18, None, "", "", 15, d(60), "M04"),
    ("FORMWORK-DAM", "Modular dam formwork", "Civil", 1800, "M2", 280, None, "", "", 60, d(100), "M04"),
    ("SHOTCRETE-TUN", "Shotcrete for diversion tunnel", "Civil", 4800, "M3", 320, None, "", "", 30, d(60), "M02"),
    ("ROCKBOLT-TUN", "Tunnel rock bolt + plate", "Civil", 18000, "EA", 42, None, "", "", 30, d(60), "M02"),
    ("CABLE-MV-DAM", "MV cable installed dam", "Cables", 18, "KM", 86000, None, "", "", 90, d(280), "M08"),
    ("CABLE-LV-DAM", "LV cable bulk dam", "Cables", 48, "KM", 22000, None, "", "", 60, d(280), "M08"),
    ("EARTH-MAT", "Generator hall earth mat", "Electrical", 8400, "M", 28, None, "", "", 60, d(280), "M08"),
    ("BATTERY-DC", "220 V DC battery bank", "Electrical", 4, "EA", 28000, None, "", "", 120, d(290), "M08"),
    ("SCADA-COMMS", "SCADA RTU + comms", "Automation", 6, "EA", 38000, None, "", "", 120, d(310), "M09"),
    ("HV-CB-220", "220 kV circuit breaker", "Electrical", 6, "EA", 320000, None, "", "", 270, d(320), "M10"),
    ("LA-220", "220 kV lightning arrester", "Electrical", 12, "EA", 18000, None, "", "", 90, d(320), "M10"),
    ("OVERHEAD-LINE", "Overhead transmission stringing", "Electrical", 18, "KM", 280000, None, "", "", 120, d(330), "M10"),
    ("FENCING-PERIM", "Perimeter security fencing", "Civil", 8400, "M", 28, None, "", "", 30, d(60), "M01"),
    ("ROAD-ACCESS-AGG", "Access road aggregate", "Civil", 28000, "T", 12, None, "", "", 15, d(45), "M01"),
    ("CULVERT-ARMCO", "Armco culvert sections", "Civil", 280, "M", 480, None, "", "", 60, d(90), "M01"),
    ("SURVEY-EQUIP", "Total station + GPS", "Tools", 6, "EA", 18000, None, "", "", 45, d(60), "M01"),
    ("CONCRETE-PUMP", "Truck-mounted concrete pump", "Civil", 4, "EA", 240000, None, "", "", 90, d(90), "M02"),
    ("WELDING-CONSUM", "Welding consumables bulk", "Tools", 8, "T", 4200, None, "", "", 30, d(60), "M02"),
    ("HYDRAULIC-CYL", "Penstock hydraulic cylinder", "Hydro turbines", 12, "EA", 28000, "Voith Industrial Services", "", "", 150, d(280), "M07"),
    ("BEARING-THRUST", "Generator thrust bearing", "Generators", 2, "EA", 145000, "Bharat Heavy Electricals", "", "", 240, d(280), "M08"),
    ("BEARING-GUIDE", "Generator guide bearing", "Generators", 4, "EA", 86000, "Bharat Heavy Electricals", "", "", 240, d(280), "M08"),
    ("PUMP-DEWATER", "Powerhouse dewatering pump", "Rotating equipment", 4, "EA", 38000, None, "", "", 120, d(280), "M08"),
    ("PUMP-DRAINAGE", "Drainage pump submersible", "Rotating equipment", 6, "EA", 18000, None, "", "", 90, d(280), "M08"),
    ("EROSION-MAT", "Erosion control geotextile", "Civil", 4800, "M2", 22, None, "", "", 45, d(75), "M03"),
    ("RIPRAP-STONE", "Riprap stone bedding", "Civil", 18000, "T", 28, None, "", "", 30, d(75), "M03"),
    ("REINFORCED-STEEL", "Reinforced steel mass concrete", "Civil", 8400, "T", 720, None, "", "", 45, d(90), "M04"),
    ("BLAST-MAT", "Blasting matting", "Civil", 240, "EA", 1800, None, "", "", 60, d(60), "M02"),
    ("TRANS-LINE-TWR", "Transmission tower steel", "Electrical", 48, "EA", 86000, None, "", "", 240, d(320), "M10"),
    ("INSULATOR-DISC", "Disc insulator string", "Electrical", 480, "EA", 380, None, "", "", 150, d(320), "M10"),
    ("CONDUCTOR-ACSR", "ACSR transmission conductor", "Cables", 84, "KM", 28000, None, "", "", 180, d(320), "M10"),
]


SUPPLIERS = [
    # name, category, country, lead_time, otd_pct, ppm, annual_spend, alternates, risk_flags, applies_to
    ("Helios Cast & Forge", "Forged valves", "India", 42, 88.0, 1800, 740000, 0, "single source; late NCR closure", "arcforge"),
    ("BluePeak Controls", "PLC and control panels", "Germany", 55, 93.0, 650, 1280000, 1, "port congestion", "arcforge,helios"),
    ("Copperline Metals", "Copper busbars", "Malaysia", 28, 97.0, 300, 510000, 2, "", "arcforge,helios"),
    ("Kerala Forge Works", "Forged valves", "India", 56, 91.0, 1400, 180000, 1, "capacity constrained", "arcforge"),
    ("Mitsuba Automation", "PLC and control panels", "Japan", 62, 96.0, 420, 320000, 2, "", "arcforge"),
    ("Delhi Metals Co", "Copper busbars", "India", 24, 94.0, 520, 210000, 1, "new supplier", "arcforge"),
    ("NorthCable Subsea", "Subsea cables", "Norway", 180, 92.0, 410, 4200000, 1, "weather window critical", "helios"),
    ("DeepDive ROV Systems", "ROV connectors", "UK", 84, 89.0, 820, 860000, 0, "single source; small batch", "helios"),
    ("Coastline Marine Coatings", "Marine epoxy", "Netherlands", 35, 95.0, 280, 240000, 2, "", "helios"),
    ("Aberdeen Steel Fab", "Offshore structural", "UK", 140, 86.0, 1100, 2900000, 1, "welder shortage", "helios"),
    ("Andritz Hydro", "Hydro turbines", "Austria", 420, 94.0, 180, 6800000, 0, "single source; long lead", "northwind"),
    ("ThyssenKrupp Specialty Plate", "Alloy plate", "Germany", 110, 92.0, 320, 1500000, 2, "", "northwind"),
    ("Bharat Heavy Electricals", "Generators", "India", 300, 88.0, 520, 4100000, 1, "capacity constrained", "northwind"),
    ("Voith Industrial Services", "Penstock fabrication", "Germany", 240, 91.0, 410, 2200000, 1, "transport over-dimensional", "northwind"),
    ("MeridianAlloys (test)", "Alloy plate", "India", 90, 89.0, 720, 0, 2, "new supplier", "arcforge,northwind"),
    ("ApexInstruments (test)", "Instrumentation", "USA", 90, 95.0, 200, 0, 3, "", "arcforge,helios,northwind"),
    ("StratoLogistics (test)", "Freight forwarding", "Singapore", 14, 92.0, 0, 0, 5, "", "arcforge,helios,northwind"),
    ("Trident Pumps (test)", "Rotating equipment", "China", 150, 87.0, 980, 0, 2, "trade tariffs", "arcforge,helios,northwind"),
]


# Test scenarios — covers every major feature with a click-path.
SCENARIOS = [
    # id, area, persona, steps, expected, priority
    ("T01", "Auth · login", "any", "Open /login → pick any persona → land on /overview", "Redirected to /overview; topbar shows correct tenant + role", "P1"),
    ("T02", "Auth · expired token", "arcforge-admin-01", "Edit sessionStorage.sct.auth.token.v1 to invalid → reload", "Frontend clears token + redirects back to /login (no 'Token expired' wall)", "P1"),
    ("T03", "Auth · cross-tenant 404", "arcforge-buyer-01", "GET /api/projects/HYD-MAHADEV-220 with arcforge token", "HTTP 404 (helios/northwind data not visible)", "P1"),
    ("T04", "Portfolio · overview", "any-admin", "Open /overview", "Donut shows average completion %, spend rollup, buckets, schedule + activity panels", "P1"),
    ("T05", "Portfolio · tenant differs", "all three admins", "Login each, compare /overview", "Each tenant shows distinct project counts, budgets, sectors", "P1"),
    ("T06", "Portfolio · cache TTL", "arcforge-admin-01", "Call /api/portfolio/summary twice within 10s, then after 15s", "Server-side cache returns same payload twice quickly, then refreshes", "P2"),
    ("T07", "Command palette · open", "any", "Press ⌘K (Mac) or Ctrl+K (other)", "Modal opens centred; search input focused; lists default items", "P1"),
    ("T08", "Command palette · search vendor", "arcforge-buyer-01", "⌘K → type 'helios'", "Helios Cast & Forge vendor row appears; Enter navigates to vendor page", "P1"),
    ("T09", "Command palette · search BOM", "northwind-buyer-01", "⌘K → type 'runner'", "Francis runner BOM rows show; click navigates to project with bom=…", "P1"),
    ("T10", "Projects · list + cards", "any", "Open /projects", "4 cards per tenant; each shows completion bar + next milestone + In Xd; in-flight project ~55% (amber)", "P1"),
    ("T11", "Projects · skeleton + animation", "any", "Hard reload /projects with throttled network (DevTools Slow 3G)", "Shimmer skeletons render first, then cards fade-up", "P2"),
    ("T12", "Projects · hover prefetch", "any", "Hover a project card before clicking", "Detail page loads near-instant (prefetched JSON + Next route)", "P2"),
    ("T13", "Project detail · completion breakdown", "any", "Open an in-flight project", "Top panel shows blended % + Milestones / BOM delivered / Spend committed columns", "P1"),
    ("T14", "Project detail · milestones timeline", "any", "Scroll milestones section", "Past milestones marked green; future ones outlined; date + 'in Xd' / 'Xd ago'", "P2"),
    ("T15", "BOM · list + status chips", "any", "Project detail → BOM tab", "Table shows code, qty, supplier, lead, need by, milestone, spec, status (planned / spec_missing / ordered / delivered)", "P1"),
    ("T16", "BOM · search + status filter", "any", "Type in search; pick 'spec_missing' from status dropdown", "Rows filter live; counter '5 of 9' updates", "P2"),
    ("T17", "BOM · row detail modal", "any", "Click any BOM row", "Modal opens with full detail + spec/drawing refs + Create PR shortcut", "P1"),
    ("T18", "BOM · upload CSV", "arcforge-head-01", "Project detail → BOM tab → expand 'CSV format' → upload test-data/bom-arcforge-PRJ-RB-660.csv", "rows_parsed > 0, rows_accepted == lines; new BOM items appear; audit event emitted", "P1"),
    ("T19", "Procurement plan · generate", "any", "Project detail → Procurement Plan tab", "Plan packages grouped by milestone; long-lead + missing-spec flags listed; assumptions visible", "P1"),
    ("T20", "PR · create from BOM", "arcforge-buyer-01", "BOM row → Create PR → fill buyer + need by → submit", "POST /api/prs returns PR; row shows status updated", "P1"),
    ("T21", "PR · suggested vendors", "arcforge-buyer-01", "Open PR detail → check suggested vendors panel", "Top suppliers from supplier list shown, filtered to category", "P2"),
    ("T22", "RFQ · issue from PR", "arcforge-buyer-01", "PR detail → Issue RFQ → pick 3 vendors → submit", "RFQ created; PR status changes to rfq_issued", "P1"),
    ("T23", "Quote · receive", "arcforge-buyer-01", "RFQ detail → Add Quote → enter vendor/price/lead/incoterm", "Quote saved; PR status → quoted", "P1"),
    ("T24", "TBE · score quotes", "arcforge-head-01", "RFQ → TBE tab → set criteria + auto-evaluate", "Technical scores blended with commercial composite; recommendation surfaces", "P2"),
    ("T25", "Award · single-source", "arcforge-head-01", "RFQ with 1 vendor only → award", "Award created with rationale (LLM or deterministic); PO drafted", "P1"),
    ("T26", "Award · multi-vendor", "arcforge-head-01", "RFQ with 3 quotes → award winner", "PO created; vendor + value match award; audit chain shows pr_created → rfq_issued → quote_received → awarded → po_drafted", "P1"),
    ("T27", "PO · timeline", "any", "PO detail → Timeline tab", "All upstream events ordered + linked back to PR / RFQ / Quote", "P2"),
    ("T28", "Vendors · intel list", "any", "Open /vendors", "Scorecard table with grades + composite scores + flags", "P1"),
    ("T29", "Vendors · scorecard radar", "any", "Vendor detail page", "6-axis radar (delivery/quality/price/responsiveness/claims/risk), alternates ranked", "P2"),
    ("T30", "Vendors · category concentration", "any", "/vendors → concentration tab", "Top vendor share % per category; single_source flag set when n=1", "P2"),
    ("T31", "Expediting · queue", "any", "Open /expediting", "Items sorted by slip probability; urgency badges OK/Watch/Nudge/Escalate; value at risk", "P1"),
    ("T32", "Expediting · follow-up email", "arcforge-expeditor-01", "Click 'Draft follow-up' on a delayed PO → pick tone urgent", "Modal shows body + requested docs; Save logs audit event", "P2"),
    ("T33", "Logistics · shipment list", "any", "Open /logistics", "Shipments sorted by bottleneck first, then slack; bottleneck badges visible", "P1"),
    ("T34", "Logistics · add event", "arcforge-expeditor-01", "Open shipment → add event with stage in_transit + location", "New event appended; audit trail records stage_advanced", "P2"),
    ("T35", "Logistics · mode recommender", "any", "Shipment detail → Recommend Mode", "Returns recommended mode + cost multiplier + transit estimate; rationale cited", "P2"),
    ("T36", "Commercial · roll-up", "any-admin", "Open /commercial", "Per-project budget vs quoted vs awarded; top savings + over-budget rows", "P1"),
    ("T37", "Commercial · scope to tenant", "all three admins", "Compare /commercial across tenants", "Each tenant sees distinct project list and totals", "P1"),
    ("T38", "Simulations · vendor slip", "arcforge-head-01", "/simulate → vendor_slip_2w → target Helios Cast & Forge", "Result shows cost delta + schedule delta + milestone impacts + mitigations", "P2"),
    ("T39", "Simulations · customs hold", "any", "/simulate → customs_hold → target a sourcing PO", "Result + LLM narrative (or deterministic fallback)", "P2"),
    ("T40", "Simulations · alt vendor", "arcforge-head-01", "alt_vendor target Helios → alternate Kerala Forge Works", "Result with delta vs current; mitigation list", "P2"),
    ("T41", "Weekly plan · build", "any-admin", "Open /weekly-plan", "P1/P2/P3 items grouped; AI narrative (Grok) or deterministic; KPI snapshot", "P1"),
    ("T42", "Agent chat · tool call", "arcforge-head-01", "/agent → ask 'Show me open PRs for Riverbank'", "Reply references open PRs (tool _tool_open_prs); source: 'grok' or 'deterministic'", "P2"),
    ("T43", "Risks · mitigations", "arcforge-head-01", "/risks → click 'Mitigations' on top risk", "3 concrete mitigations + audit log; source flag visible", "P2"),
    ("T44", "Audit · entity trace", "any", "Open /audit → click any pr_no", "Full chain from BOM → PR → RFQ → Quote → Award → PO renders", "P1"),
    ("T45", "Audit · pivots + CSV export", "any", "Open /audit → pivots → Materials → 'Export CSV'", "CSV downloads with all filtered events; filename includes timestamp", "P2"),
    ("T46", "SAP CPI · submit PR", "arcforge-head-01", "PR detail → 'Submit to SAP'", "sap_status transitions submitting → synced (or mock); sap_pr_no populated", "P2"),
    ("T47", "SAP CPI · webhook event", "external", "POST /api/integrations/sap/event with mock payload", "sap_status updated on matching PR/PO; event recorded", "P3"),
    ("T48", "Integrations · health", "any-admin", "/integrations → SAP card", "Health response shows mode + last event time", "P3"),
    ("T49", "RBAC · viewer write denied", "arcforge-viewer-01", "Try POST /api/prs as viewer", "HTTP 403 (require_perm('pr','create') triggers)", "P1"),
    ("T50", "RBAC · admin tenant switch", "arcforge-admin-01", "GET /api/tenants then switch token via X-Tenant-Override header", "Admin can read other tenants; non-admin cannot", "P2"),
    # ---- Edge-case CSV uploads (added with backend extension to accept status + parent_item_id) ----
    ("T51", "CSV upload · error reporting", "arcforge-head-01", "Upload bom-edge-malformed.csv to PRJ-AF-CCGT", "rows_rejected ≥ 4 (empty code, bad qty, bad cost, bad date, bad lead); 'errors' list cites row numbers + reasons; the 4 OK rows are accepted; the bad-status row gets a warning and falls back to default", "P1"),
    ("T52", "CSV upload · minimal columns", "arcforge-head-01", "Upload bom-edge-minimal.csv to PRJ-AF-SUB (only code/description/quantity)", "All 5 rows accepted; uom defaults to 'EA'; status = 'spec_missing'; no errors", "P2"),
    ("T53", "CSV upload · status column honored", "arcforge-head-01", "Upload bom-status-mix-PRJ-RB-660.csv to PRJ-RB-660", "9 rows accepted with delivered/ordered/planned statuses preserved; Riverbank's completion % jumps from ~0% to ~30-40% on next /overview refresh (cache TTL 10s)", "P1"),
    ("T54", "CSV upload · status mix moves portfolio", "helios-head-01", "Upload bom-status-mix-PRJ-HE-FPSO.csv to PRJ-HE-FPSO", "FPSO completion bucket moves from 'Planning (0%)' to 'In progress (25-70%)'; portfolio donut average climbs", "P1"),
    ("T55", "CSV upload · status mix northwind", "northwind-head-01", "Upload bom-status-mix-PRJ-NW-STEEL.csv to PRJ-NW-STEEL", "Polaris Steel project completion ~30-40%; spend_committed_pct on /api/projects/{id}/progress shows the delivered + ordered value", "P1"),
    ("T56", "CSV upload · unknown status warning", "any-head", "Upload bom-edge-malformed.csv → row EDGE-BADSTATUS", "errors list contains 'unknown status \"super_dispatched\" — using default'; row still accepted with derived status", "P2"),
    ("T57", "CSV upload · case-insensitive status", "any-head", "Upload bom-edge-malformed.csv → row EDGE-OK-04", "Status 'ORDERED' normalised to lower-case 'ordered'; row accepted", "P2"),
    ("T58", "CSV upload · cross-tenant denial", "arcforge-buyer-01", "Try uploading bom-status-mix-PRJ-NW-STEEL.csv to PRJ-NW-STEEL with arcforge token", "Returns 404 (project_id appears 'unknown' from arcforge tenant's perspective)", "P1"),
]


# Edge-case + status-mix CSVs catalog (auxiliary uploads).
AUX_FILES = [
    # filename, purpose, target_project, sample_size, expected_behaviour
    ("bom-edge-malformed.csv", "Error-path probe", "any greenfield (e.g. PRJ-AF-CCGT)",
     "12 rows, 6 deliberately bad", "rows_rejected count + per-row error messages"),
    ("bom-edge-minimal.csv", "Required-fields-only path", "any greenfield",
     "5 rows, only code/description/quantity", "All accepted, uom defaults to EA, status = spec_missing"),
    ("bom-status-mix-PRJ-RB-660.csv", "Status-mix for Riverbank", "PRJ-RB-660 (arcforge)",
     "9 rows: 5 delivered + 1 ordered + 3 planned", "Drives Riverbank to in-flight"),
    ("bom-status-mix-PRJ-HE-FPSO.csv", "Status-mix for Hawthorn FPSO", "PRJ-HE-FPSO (helios)",
     "9 rows: 5 delivered + 1 ordered + 3 planned", "Drives FPSO to in-flight"),
    ("bom-status-mix-PRJ-NW-STEEL.csv", "Status-mix for Polaris Steel Mill", "PRJ-NW-STEEL (northwind)",
     "9 rows: 5 delivered + 1 ordered + 3 planned", "Drives steel mill to in-flight"),
]


# API reference — public surface worth knowing.
APIS = [
    ("POST", "/api/auth/login", "no", "Issue JWT for picked persona", "{ \"user_id\": \"arcforge-admin-01\" }"),
    ("GET", "/api/auth/me", "yes", "Current user + permissions", ""),
    ("GET", "/api/auth/personas", "no", "All seeded personas across tenants", ""),
    ("GET", "/api/portfolio/summary", "yes", "Tenant cockpit aggregation (10s cache)", ""),
    ("GET", "/api/search/index", "yes", "Cmd+K palette index", ""),
    ("GET", "/api/projects", "yes", "List tenant projects", ""),
    ("GET", "/api/projects/progress", "yes", "Bulk completion % across projects", ""),
    ("GET", "/api/projects/{id}", "yes", "Project detail", ""),
    ("GET", "/api/projects/{id}/progress", "yes", "Single project completion breakdown", ""),
    ("GET", "/api/projects/{id}/bom", "yes", "BOM lines for project", ""),
    ("POST", "/api/projects/{id}/bom/upload", "yes", "Upload BOM CSV (multipart file)", ""),
    ("GET", "/api/projects/{id}/procurement-plan", "yes", "Build packages + long-lead + missing-spec flags", ""),
    ("POST", "/api/projects/{id}/bom/autofill", "yes", "AI fill missing supplier/category", ""),
    ("POST", "/api/projects/{id}/bom/{bom_id}/spec-request", "yes", "AI draft spec-request email", ""),
    ("GET", "/api/prs", "yes", "List PRs (tenant-scoped)", ""),
    ("POST", "/api/prs", "yes", "Create PR", "{ project_id, bom_item_id, ... }"),
    ("GET", "/api/prs/{pr_no}", "yes", "PR detail", ""),
    ("GET", "/api/prs/{pr_no}/suggested-vendors", "yes", "Vendor suggestions for PR", ""),
    ("POST", "/api/prs/{pr_no}/submit-to-sap", "yes", "Submit to SAP CPI", ""),
    ("GET", "/api/rfqs", "yes", "List RFQs", ""),
    ("POST", "/api/rfqs", "yes", "Issue RFQ from PR", "{ pr_no, vendors[], due_in_days, notes }"),
    ("GET", "/api/rfqs/{rfq_no}/quotes", "yes", "Quotes received", ""),
    ("POST", "/api/rfqs/{rfq_no}/quotes", "yes", "Receive quote", "{ vendor, unit_price_usd, lead_time_days, ... }"),
    ("GET", "/api/rfqs/{rfq_no}/compare", "yes", "Quote comparison + winner", ""),
    ("POST", "/api/rfqs/{rfq_no}/award", "yes", "Award RFQ + draft PO", "{ quote_id, rationale? }"),
    ("GET", "/api/sourcing-pos", "yes", "List sourcing POs", ""),
    ("GET", "/api/sourcing-pos/{po_no}/timeline", "yes", "Full audit chain for PO", ""),
    ("POST", "/api/sourcing-pos/{po_no}/submit-to-sap", "yes", "Submit PO to SAP CPI", ""),
    ("GET", "/api/vendors/intel", "yes", "Vendor scorecard list", ""),
    ("GET", "/api/vendors/intel/{name}", "yes", "Single scorecard + alternates", ""),
    ("GET", "/api/vendors/intel/{name}/briefing", "yes", "AI risk briefing (Grok or deterministic)", ""),
    ("GET", "/api/vendors/concentration", "yes", "Category concentration", ""),
    ("GET", "/api/expediting/queue", "yes", "Expedite queue with urgency", ""),
    ("POST", "/api/expediting/{po}/draft-followup", "yes", "AI draft tone-aware follow-up email", ""),
    ("GET", "/api/logistics/shipments", "yes", "Shipment list", ""),
    ("POST", "/api/logistics/shipments/{po}/events", "yes", "Add shipment event", ""),
    ("GET", "/api/commercial/summary", "yes", "Budget vs quoted vs awarded roll-up", ""),
    ("POST", "/api/risk/simulate", "yes", "Run scenario (vendor_slip / customs_hold / alt_vendor)", ""),
    ("GET", "/api/weekly-plan", "yes", "Weekly priority plan + AI narrative", ""),
    ("POST", "/api/chat", "yes", "Agent dispatch with tool calling", "{ message, history? }"),
    ("POST", "/api/risks/mitigations", "yes", "AI mitigations for a risk record", ""),
    ("POST", "/api/explain", "yes", "Explain any entity (po/vendor/risk/project/rfq/pr)", ""),
    ("GET", "/api/audit", "yes", "Paginated audit events", ""),
    ("GET", "/api/audit/export.csv", "yes", "Download filtered audit log as CSV", ""),
    ("GET", "/healthz", "no", "Liveness probe", ""),
    ("GET", "/readyz", "no", "Readiness (snapshot status)", ""),
    ("POST", "/api/admin/snapshot", "yes", "Force in-memory snapshot to disk", ""),
]


# ---- write workbook ---------------------------------------------------------

def build_workbook(out_path: Path) -> None:
    wb = Workbook()
    # README
    ws = wb.active
    ws.title = "README"
    ws.cell(row=1, column=1, value="SUPPLY CHAIN CONTROL TOWER").font = EYEBROW_FONT
    ws.cell(row=2, column=1, value="Test Data Pack").font = TITLE_FONT
    intro = [
        ("Purpose",
         "End-to-end test data + scenarios for every feature in the Control Tower. "
         "Use the persona table to log in, the BOM CSVs to populate projects, the suppliers list as reference, "
         "and the scenarios sheet as a click-by-click QA script."),
        ("Live environments",
         "Local: https://localhost/  ·  Fly: https://scm-towerx.fly.dev/"),
        ("How to use",
         "1. Open the 'Personas' sheet — pick any user_id to sign in.\n"
         "2. Open the 'Test Scenarios' sheet — work through P1 first.\n"
         "3. To populate a project's BOM, use the matching 'bom-<tenant>-<project>.csv' file via the project's BOM tab → CSV upload.\n"
         "4. To exercise edge cases + watch completion respond, see the 'Edge / Status CSVs' sheet (bom-edge-* and bom-status-mix-*).\n"
         "5. The 'Suppliers' and 'BOM Library' sheets are reference data for manual entry / API calls.\n"
         "6. The 'API Reference' sheet documents every endpoint to hit directly with curl / Postman."),
        ("Tenants",
         "arcforge (Power Systems EPC)  ·  helios (Offshore Engineering)  ·  northwind (Heavy Engineering). "
         "All data is tenant-scoped — cross-tenant reads return 404."),
        ("Auth",
         "POST /api/auth/login with { \"user_id\": \"<persona>\" } returns a JWT. "
         "Send as Authorization: Bearer <token>. 8h TTL. JWT secret is set via the JWT_SECRET env var in prod."),
    ]
    r = 4
    for label, body in intro:
        ws.cell(row=r, column=1, value=label).font = SECTION_FONT
        ws.cell(row=r, column=2, value=body).font = BODY_FONT
        ws.cell(row=r, column=2).alignment = WRAP
        ws.row_dimensions[r].height = max(36, 18 * (1 + body.count("\n")))
        r += 1
    set_widths(ws, [22, 110])

    # Personas
    ws = wb.create_sheet("Personas")
    write_table(
        ws,
        eyebrow="AUTH",
        title="Personas — pick any user_id to sign in",
        headers=["user_id", "Display name", "Email", "Tenant", "Role", "What this persona can do"],
        rows=personas(),
        widths=[26, 28, 32, 30, 18, 70],
    )

    # Projects
    ws = wb.create_sheet("Projects")
    write_table(
        ws,
        eyebrow="PORTFOLIO",
        title="Projects — 4 per tenant, 1 in-execution each",
        headers=["Tenant", "project_id", "Name", "Client", "Site", "Sector",
                 "Expected completion (today)", "Seeded BOM lines"],
        rows=[(p[0], *p[1:]) for p in PROJECTS],
        widths=[14, 22, 46, 36, 32, 26, 28, 18],
    )

    # BOM Library — split by sheet per tenant for clarity
    def write_bom(sheet_name, rows, anchor_project):
        ws = wb.create_sheet(sheet_name)
        # Build status from spec_doc_id (spec_missing if blank)
        rendered = []
        for (code, desc, cat, qty, uom, cost, supplier, spec, drawing, lead, need, ms) in rows:
            status = "spec_missing" if not spec else "planned"
            rendered.append([anchor_project, code, desc, cat, qty, uom, cost,
                             supplier or "", spec, drawing, lead, need, ms, status])
        last = write_table(
            ws,
            eyebrow="BOM",
            title=f"{sheet_name} — uploadable to {anchor_project}",
            headers=[
                "project_id", "code", "description", "category", "quantity", "uom",
                "unit_cost_usd", "supplier_name", "spec_doc_id", "drawing_id",
                "long_lead_days", "planned_need_date", "milestone_code", "status",
            ],
            rows=rendered,
            widths=[18, 22, 36, 22, 10, 8, 14, 26, 18, 16, 14, 16, 14, 14],
        )
        # Totals row (formula!)
        header_row_idx = 4  # eyebrow + title + blank + header
        first_data = header_row_idx + 1
        last_data = first_data + len(rendered) - 1
        totals_row = last_data + 2
        ws.cell(row=totals_row, column=1, value="TOTAL LINES").font = SECTION_FONT
        ws.cell(row=totals_row, column=5,
                value=f"=COUNTA(B{first_data}:B{last_data})")
        ws.cell(row=totals_row, column=7,
                value=f"=SUMPRODUCT(E{first_data}:E{last_data},G{first_data}:G{last_data})")
        ws.cell(row=totals_row, column=7).number_format = "$#,##0"

    write_bom("BOM-Arcforge", BOM_ARCFORGE, "PRJ-RB-660")
    write_bom("BOM-Helios", BOM_HELIOS, "PRJ-NS-OSS")
    write_bom("BOM-Northwind", BOM_NORTHWIND, "HYD-MAHADEV-220")

    # Suppliers
    ws = wb.create_sheet("Suppliers")
    write_table(
        ws,
        eyebrow="VENDORS",
        title="Suppliers — reference list (rows tagged 'test' are net-new)",
        headers=[
            "name", "category", "country", "lead_time_days", "on_time_delivery_pct",
            "quality_ppm", "annual_spend_usd", "approved_alternatives", "risk_flags",
            "Applies to tenants",
        ],
        rows=SUPPLIERS,
        widths=[34, 26, 18, 14, 18, 14, 18, 18, 28, 26],
    )
    # Number format spend
    for r in range(5, 5 + len(SUPPLIERS)):
        ws.cell(row=r, column=7).number_format = "$#,##0;($#,##0);-"
        ws.cell(row=r, column=5).number_format = "0.0\"%\""

    # Test Scenarios
    ws = wb.create_sheet("Test Scenarios")
    write_table(
        ws,
        eyebrow="QA",
        title="Test Scenarios — work P1 first, then P2/P3",
        headers=["ID", "Feature area", "Persona", "Steps", "Expected outcome", "Priority", "Pass?"],
        rows=[(s[0], s[1], s[2], s[3], s[4], s[5], "") for s in SCENARIOS],
        widths=[8, 30, 26, 60, 60, 12, 10],
    )

    # Feature Coverage — count by area
    ws = wb.create_sheet("Coverage Matrix")
    write_table(
        ws,
        eyebrow="MATRIX",
        title="Feature coverage — counts by area (formula-driven)",
        headers=["Area", "P1 count", "P2 count", "P3 count", "Total"],
        rows=[],
        widths=[34, 14, 14, 14, 14],
    )
    areas = sorted({s[1].split(" · ")[0] for s in SCENARIOS})
    header_row = 4
    for i, area in enumerate(areas):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=area).font = BODY_FONT
        ws.cell(row=r, column=2, value=f'=COUNTIFS(\'Test Scenarios\'!B:B,A{r}&"*",\'Test Scenarios\'!F:F,"P1")')
        ws.cell(row=r, column=3, value=f'=COUNTIFS(\'Test Scenarios\'!B:B,A{r}&"*",\'Test Scenarios\'!F:F,"P2")')
        ws.cell(row=r, column=4, value=f'=COUNTIFS(\'Test Scenarios\'!B:B,A{r}&"*",\'Test Scenarios\'!F:F,"P3")')
        ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})")
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = BOX
            ws.cell(row=r, column=col).font = BODY_FONT
    # Totals
    totals = header_row + 1 + len(areas) + 1
    ws.cell(row=totals, column=1, value="TOTAL").font = SECTION_FONT
    for col in range(2, 6):
        letter = get_column_letter(col)
        ws.cell(row=totals, column=col,
                value=f"=SUM({letter}{header_row+1}:{letter}{header_row+len(areas)})").font = SECTION_FONT

    # Edge-case + status-mix CSV catalogue
    ws = wb.create_sheet("Edge + Status CSVs")
    write_table(
        ws,
        eyebrow="EDGE CASES",
        title="Auxiliary CSV files for robustness + completion testing",
        headers=["File", "Purpose", "Target project", "Contents", "Expected behaviour"],
        rows=AUX_FILES,
        widths=[42, 30, 32, 38, 56],
    )
    # Add a quick-link note row
    last_row = ws.max_row + 2
    ws.cell(row=last_row, column=1, value="HOW TO LOAD").font = SECTION_FONT
    ws.cell(row=last_row + 1, column=1, value=(
        "Sign in as a *-head-01 or *-admin-01 persona for the matching tenant. "
        "Open the target project → BOM tab → expand 'CSV format' → drop the file. "
        "The backend's upload_bom_csv now accepts an optional 'status' column "
        "(spec_missing | planned | requisitioned | ordered | delivered) and an "
        "optional 'parent_item_id' column — unknown statuses warn + fall back."
    )).alignment = WRAP
    ws.cell(row=last_row + 1, column=1).font = BODY_FONT
    ws.row_dimensions[last_row + 1].height = 60

    # API Reference
    ws = wb.create_sheet("API Reference")
    write_table(
        ws,
        eyebrow="API",
        title="Public endpoints — auth via Bearer JWT unless marked 'no'",
        headers=["Method", "Endpoint", "Auth", "What it does", "Sample body"],
        rows=APIS,
        widths=[10, 56, 8, 60, 38],
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ---- write CSV files --------------------------------------------------------

def write_bom_csv(rows, anchor_project, out_path: Path) -> None:
    fieldnames = [
        "code", "description", "category", "quantity", "uom", "unit_cost_usd",
        "supplier_name", "spec_doc_id", "drawing_id", "long_lead_days",
        "planned_need_date", "milestone_code",
    ]
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for (code, desc, cat, qty, uom, cost, supplier, spec, drawing, lead, need, ms) in rows:
            w.writerow({
                "code": code, "description": desc, "category": cat,
                "quantity": qty, "uom": uom, "unit_cost_usd": cost,
                "supplier_name": supplier or "", "spec_doc_id": spec,
                "drawing_id": drawing, "long_lead_days": lead,
                "planned_need_date": need, "milestone_code": ms,
            })


def main() -> None:
    build_workbook(HERE / "control-tower-test-pack.xlsx")
    write_bom_csv(BOM_ARCFORGE, "PRJ-RB-660", HERE / "bom-arcforge-PRJ-RB-660.csv")
    write_bom_csv(BOM_HELIOS, "PRJ-NS-OSS", HERE / "bom-helios-PRJ-NS-OSS.csv")
    write_bom_csv(BOM_NORTHWIND, "HYD-MAHADEV-220", HERE / "bom-northwind-HYD-MAHADEV-220.csv")
    print("Wrote:")
    for p in sorted(HERE.glob("*")):
        if p.name.startswith("_"):
            continue
        print(f"  {p.name}  ({p.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
